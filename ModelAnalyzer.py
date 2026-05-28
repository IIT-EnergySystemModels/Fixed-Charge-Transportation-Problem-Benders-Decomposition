from gurobipy import read, Model, GRB
import pandas as pd
from typing import Union
import os
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter


# ==============================
# CONFIGURATION
# ==============================
CASE_NAME = 'ModelName_with_symbolic_labels' # name of the .lp file without extension
FILE_PATH = ''  # directory of .lp; '' → cwd

VARIABLES_SUFFIX = '_Variables.csv'
CONSTRAINTS_SUFFIX = '_Constraints.csv'
VAR_STEMS_SUFFIX = '_VarStems.csv'
CON_STEMS_SUFFIX = '_ConStems.csv'
SUMMARY_SUFFIX = '_Summary.xlsx'

EPSILON = 1e-6

def get_text_before_parenthesis(text: str) -> str:
    idx = text.find('(')
    return text if idx == -1 else text[:idx]


def deduplicate_by_key(data: list[dict], key: str) -> list[dict]:
    seen = {}
    for item in data:
        seen[item[key]] = item
    return list(seen.values())


def auto_fit_columns(ws):
    '''Resize Excel columns to fit content.'''
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2


def apply_percentage_and_colors(ws, col_names: list[str]):
    '''
    Apply percentage formatting and a green-yellow-red color scale
    to given header-named columns in an Excel worksheet.
    '''
    # Map headers to column numbers
    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}

    for name in col_names:
        if name not in headers:
            continue

        col_idx = headers[name]
        col_letter = get_column_letter(col_idx)
        col_range = f'{col_letter}2:{col_letter}{ws.max_row}'  # skip header

        # Apply percentage formatting to all cells in this column
        for row in range(2, ws.max_row + 1):
            cell = ws[f'{col_letter}{row}']
            if cell.value is None:
                continue
            cell.number_format = '0.00%'

        # Apply a 3-color gradient: red → yellow → green
        color_scale_rule = ColorScaleRule(
            start_type='num', start_value=0, start_color='FFC7CE',  # soft green
            mid_type='num', mid_value=0.5, mid_color='FFEB9C',  # soft yellow
            end_type='num', end_value=1, end_color='C6EFCE'  # soft red/pink
        )
        ws.conditional_formatting.add(col_range, color_scale_rule)


def write_variables_constraints(dir_name: str, case_name: str, model: Model, presolved_model: Model) -> tuple[pd.DataFrame, pd.DataFrame]:
    # ---------- VARIABLES ----------
    vars_original = sorted(model.getVars(), key=lambda v: v.VarName)
    vars_presolved = sorted(presolved_model.getVars(), key=lambda v: v.VarName)

    df_vars = pd.DataFrame(
        [
            {
                'Name': v.VarName,
                'Stem': get_text_before_parenthesis(v.VarName),
                'Type': v.VType,
                'LB': v.LB,
                'Level': getattr(v, 'X', None),
                'UB': v.UB,
                'RC': getattr(v, 'RC', None),
                'Range': v.UB - v.LB,
            }
            for v in vars_original
        ]
    )

    df_vars_ps = pd.DataFrame(
        [
            {
                'Name': v.VarName,
                'Type_ps': v.VType,
                'LB_ps': v.LB,
                'Level_ps': getattr(v, 'X', None),
                'UB_ps': v.UB,
                'RC_ps': getattr(v, 'RC', None),
                'Range_ps': v.UB - v.LB,
            }
            for v in vars_presolved
        ]
    )

    df_vars = df_vars.merge(df_vars_ps, on='Name', how='outer')

    # Reduction column
    def calculate_avgrange(r):
        if pd.isna(r['Type_ps']):
            return pd.NA
        return r['Range']
    df_vars['Range'] = df_vars.apply(calculate_avgrange, axis=1)

    # Reduction column
    def calculate_reduction(r):
        if pd.isna(r['Type_ps']):
            return pd.NA
        if r['Range'] != 0:
            return 1 - r['Range_ps'] / r['Range']
        return pd.NA
    df_vars['Reduction'] = df_vars.apply(calculate_reduction, axis=1)

    # Removed column
    df_vars['Removed'] = df_vars['Type_ps'].isna().astype(int)

    # Superfluous classification (kept for compatibility)
    def classify_var(r):
        if pd.isna(r['Type_ps']):
            if r['Level'] == r['LB'] or r['Level'] == r['UB']:
                return 'NB'
            return 'B'
        return ''
    # df_vars['Superfluous'] = df_vars.apply(classify_var, axis=1)

    df_vars['Zero'] = df_vars['Level'].apply(lambda x: 1 if pd.notna(x) and abs(x) < EPSILON else 0)
    df_vars['Zero_ps'] = df_vars['Level_ps'].apply(lambda x: 1 if pd.notna(x) and abs(x) < EPSILON else 0)

    # RC flags and diff
    df_vars['RC_flag'] = df_vars['RC'].apply(lambda x: 1 if pd.notna(x) and abs(x) > EPSILON else 0)
    df_vars['RC_ps_flag'] = df_vars['RC_ps'].apply(lambda x: 1 if pd.notna(x) and abs(x) > EPSILON else 0)

    def diff_rc(r):
        if pd.notna(r['RC']) and pd.notna(r['RC_ps']) and abs(abs(r['RC']) - abs(r['RC_ps'])) > EPSILON:
            return abs(abs(r['RC']) - abs(r['RC_ps']))
        return pd.NA
    df_vars['RC_Diff'] = df_vars.apply(diff_rc, axis=1)

    def diff_LB(r):
        if pd.notna(r['Level']) and pd.notna(r['LB']) and abs(r['Level'] - r['LB']) < EPSILON:
            return 1
        return pd.NA
    df_vars['LB_flag'] = df_vars.apply(diff_LB, axis=1)

    def diff_UB(r):
        if pd.notna(r['Level']) and pd.notna(r['UB']) and abs(r['Level'] - r['UB']) < EPSILON:
            return  1
        return pd.NA
    df_vars['UB_flag'] = df_vars.apply(diff_UB, axis=1)

    def diff_LB_ps(r):
        if pd.notna(r['Level_ps']) and pd.notna(r['LB_ps']) and abs(r['Level_ps'] - r['LB_ps']) < EPSILON:
            return 1
        return pd.NA
    df_vars['LB_ps_flag'] = df_vars.apply(diff_LB_ps, axis=1)

    def diff_UB_ps(r):
        if pd.notna(r['Level_ps']) and pd.notna(r['UB_ps']) and abs(r['Level_ps'] - r['UB_ps']) < EPSILON:
            return  1
        return pd.NA
    df_vars['UB_ps_flag'] = df_vars.apply(diff_UB_ps, axis=1)

    # Save
    print('Writing variables CSV files...')
    df_vars.to_csv(f'{dir_name}/{case_name}{VARIABLES_SUFFIX}', index=False)
    stems = deduplicate_by_key([{'Stem': get_text_before_parenthesis(v.VarName)} for v in vars_original], 'Stem')
    pd.DataFrame(stems).to_csv(f'{dir_name}/{case_name}{VAR_STEMS_SUFFIX}', index=False)

    # ---------- CONSTRAINTS ----------
    cons_original = sorted(model.getConstrs(), key=lambda c: c.ConstrName)
    cons_presolved = sorted(presolved_model.getConstrs(), key=lambda c: c.ConstrName)

    df_cons = pd.DataFrame(
        [
            {
                'Name': c.ConstrName,
                'Stem': get_text_before_parenthesis(c.ConstrName),
                'Sense': c.Sense,
                'Level': c.RHS - c.Slack,
                'Bound': c.RHS,
                'Dual': getattr(c, 'Pi', None),
            }
            for c in cons_original
        ]
    )

    df_cons_ps = pd.DataFrame(
        [
            {
                'Name': c.ConstrName,
                'Sense_ps': c.Sense,
                'Level_ps': c.RHS - c.Slack,
                'Bound_ps': c.RHS,
                'Dual_ps': getattr(c, 'Pi', None),
            }
            for c in cons_presolved
        ]
    )

    df_cons = df_cons.merge(df_cons_ps, on='Name', how='outer')

    # Reduction constraint
    def calculate_reduction(r):
        if pd.isna(r['Sense_ps']) or r['Sense'] == '=':
            return pd.NA
        if r['Sense'] == '<' and r['Bound'] != 0:
            return 1 - r['Bound_ps'] / r['Bound']
        if r['Sense'] == '>' and r['Sense_ps'] == '>' and r['Bound_ps'] != 0:
            return 1 - r['Bound'] / r['Bound_ps']
        if r['Sense'] == '>' and r['Sense_ps'] == '<' and r['Bound_ps'] != 0:
            return 1 + r['Bound'] / r['Bound_ps']
        return pd.NA
    df_cons['Reduction'] = df_cons.apply(calculate_reduction, axis=1)

    # Removed
    df_cons['Removed'] = df_cons['Sense_ps'].isna().astype(int)

    # Dual flags
    df_cons['#Dual'] = df_cons['Dual'].apply(lambda x: 1 if pd.notna(x) and abs(x) > EPSILON else 0)
    df_cons['#Active'] = df_cons.apply(lambda r: 1 if (r['Sense'] == '=') or (r['#Dual'] == 1) else 0, axis=1)

    df_cons['#Dual_ps'] = df_cons['Dual_ps'].apply(lambda x: 1 if pd.notna(x) and abs(x) > EPSILON else 0)
    df_cons['#Active_ps'] = df_cons.apply(lambda r: 1 if (r['Sense_ps'] == '=') or (r['#Dual_ps'] == 1) else 0, axis=1)

    df_cons['Dual'] = pd.to_numeric(df_cons['Dual'], errors='coerce')
    df_cons['Dual_ps'] = pd.to_numeric(df_cons['Dual_ps'], errors='coerce')

    # DiffDual
    def diff_dual(r):
        if pd.notna(r['Dual']) and pd.notna(r['Dual_ps']) and abs(abs(r['Dual']) - abs(r['Dual_ps'])) > EPSILON:
            return abs(abs(r['Dual']) - abs(r['Dual_ps']))
        else:
            return pd.NA
    df_cons['#DiffDual'] = df_cons.apply(diff_dual, axis=1)

    # Superfluous/Instrumental classification (kept for compatibility)
    def classify_constr(r):
        if r['Sense'] != '=' and pd.isna(r['Sense_ps']):
            return 'Superfluous'
        if r['Sense'] == '=' and pd.isna(r['Sense_ps']) and abs(r['Dual']) < EPSILON:
            return 'Instrumental'
        else:
            return ''

    df_cons['Class'] = df_cons.apply(classify_constr, axis=1)

    # Save
    print('Writing constraints CSV files...')
    df_cons.to_csv(f'{dir_name}/{case_name}{CONSTRAINTS_SUFFIX}', index=False)
    cons_stems = deduplicate_by_key([{'Stem': get_text_before_parenthesis(c.ConstrName)} for c in cons_original], 'Stem',)
    pd.DataFrame(cons_stems).to_csv(f'{dir_name}/{case_name}{CON_STEMS_SUFFIX}', index=False)

    return df_vars, df_cons


def build_summary(dir_name: str, case_name: str, df_vars: pd.DataFrame, df_cons: pd.DataFrame) -> None:
    '''Compute and write summary Excel with stem-level + totals block.'''

    # ----- Constraint stem-level -----
    cons_reduction = df_cons.groupby('Stem')['Reduction'].mean().rename('%BoundReduction')
    cons_R = df_cons.groupby('Stem')['Name'].count().rename('R')
    cons_R_ps = df_cons.groupby('Stem')['Bound_ps'].apply(lambda s: s.notna().sum()).rename('R_ps')

    cons_summary = pd.concat([cons_reduction, cons_R, cons_R_ps], axis=1).reset_index()
    cons_summary['%RemovedR'] = cons_summary.apply(lambda r: 1 - r['R_ps'] / r['R'] if r['R'] else 0, axis=1)

    cons_summary['Removed'] = cons_summary['Stem'].map(df_cons.groupby('Stem')['Removed'].sum()).fillna(0).astype(int)
    cons_summary['Bound'] = cons_summary['Stem'].map(df_cons.groupby('Stem')['Bound'].mean()).fillna(0)
    cons_summary['Bound_ps'] = cons_summary['Stem'].map(df_cons.groupby('Stem')['Bound_ps'].mean()).fillna(0)
    cons_summary['#Active'] = cons_summary['Stem'].map(df_cons.groupby('Stem')['#Active'].sum()).fillna(0).astype(int)
    cons_summary['#Dual'] = cons_summary['Stem'].map(df_cons.groupby('Stem')['#Dual'].sum()).fillna(0).astype(int)
    cons_summary['#Active_ps'] = cons_summary['Stem'].map(df_cons.groupby('Stem')['#Active_ps'].sum()).fillna(0).astype(int)
    cons_summary['#Dual_ps'] = cons_summary['Stem'].map(df_cons.groupby('Stem')['#Dual_ps'].sum()).fillna(0).astype(int)
    cons_summary['Dual'] = cons_summary['Stem'].map(df_cons.groupby('Stem')['Dual'].mean()).fillna(0)
    cons_summary['Dual_ps'] = cons_summary['Stem'].map(df_cons.groupby('Stem')['Dual_ps'].mean()).fillna(0)

    # Filter constraints where both Dual and Dual_ps exist and are different
    cons_diffdual = df_cons.dropna(subset=['Dual', 'Dual_ps']).copy()
    cons_diffdual = cons_diffdual[abs(abs(df_cons['Dual']) - abs(df_cons['Dual_ps'])) > EPSILON]

    # Count of differing duals per stem
    cons_summary['#DiffDual'] = cons_summary['Stem'].map(cons_diffdual.groupby('Stem')['Dual_ps'].count()).fillna(0).astype(int)

    # Average difference for those differing duals
    cons_diffdual['DiffValue'] = abs(abs(df_cons['Dual']) - abs(df_cons['Dual_ps']))
    cons_summary['DiffDual'] = cons_summary['Stem'].map(cons_diffdual.groupby('Stem')['DiffValue'].mean()).fillna(0)

    # Special column
    # cons_summary['Special'] = cons_summary.apply(
    #     lambda r: r['R_ps'] - r['Active_ps'] if r['R_ps'] != r['Active_ps'] else pd.NA, axis=1)

    # ----- Variable stem-level -----
    var_reduction = df_vars.groupby('Stem')['Reduction'].mean().rename('%RangeReduction')
    var_C = df_vars.groupby('Stem')['Name'].count().rename('C')
    var_C_ps = df_vars.groupby('Stem')['Range_ps'].apply(lambda s: s.notna().sum()).rename('C_ps')

    var_summary = pd.concat([var_reduction, var_C, var_C_ps], axis=1).reset_index()
    var_summary['%RemovedC'] = var_summary.apply(lambda r: 1 - r['C_ps'] / r['C'] if r['C'] else 0, axis=1)

    # Removed variables
    var_summary['Removed'] = var_summary['Stem'].map(df_vars.groupby('Stem')['Removed'].sum()).fillna(0).astype(int)

    var_summary['RC'] = var_summary['Stem'].map(df_vars.groupby('Stem')['RC'].mean()).fillna(0)
    var_summary['RC_ps'] = var_summary['Stem'].map(df_vars.groupby('Stem')['RC_ps'].mean()).fillna(0)

    var_summary['Range'] = var_summary['Stem'].map(df_vars.groupby('Stem')['Range'].mean()).fillna(0)
    var_summary['Range_ps'] = var_summary['Stem'].map(df_vars.groupby('Stem')['Range_ps'].mean()).fillna(0)

    # RC flags counts
    var_summary['#RC'] = var_summary['Stem'].map(df_vars.groupby('Stem')['RC_flag'].sum()).fillna(0).astype(int)
    var_summary['#RC_ps'] = var_summary['Stem'].map(df_vars.groupby('Stem')['RC_ps_flag'].sum()).fillna(0).astype(int)

    # LB and UB flags counts
    var_summary['#LB'] = var_summary['Stem'].map(df_vars.groupby('Stem')['LB_flag'].sum()).fillna(0).astype(int)
    var_summary['#LB_ps'] = var_summary['Stem'].map(df_vars.groupby('Stem')['LB_ps_flag'].sum()).fillna(0).astype(int)
    var_summary['#UB'] = var_summary['Stem'].map(df_vars.groupby('Stem')['UB_flag'].sum()).fillna(0).astype(int)
    var_summary['#UB_ps'] = var_summary['Stem'].map(df_vars.groupby('Stem')['UB_ps_flag'].sum()).fillna(0).astype(int)

    # zero counts
    var_summary['%Zero'] = var_summary['Stem'].map(df_vars.groupby('Stem')['Zero'].mean()).fillna(0)
    var_summary['%Zero_ps'] = var_summary['Stem'].map(df_vars.groupby('Stem')['Zero_ps'].mean()).fillna(0)

    # Filter variables where both RC and RC_ps exist and differ
    var_diff_rc = df_vars.dropna(subset=['RC', 'RC_ps']).copy()
    var_diff_rc = var_diff_rc[abs(abs(var_diff_rc['RC']) - abs(var_diff_rc['RC_ps'])) > EPSILON]

    # Count of differing RCs per stem
    var_summary['#DiffRC'] = var_summary['Stem'].map(df_vars.groupby('Stem')['RC'].count()).fillna(0).astype(int)

    # Average difference for those differing RCs
    var_diff_rc['DiffValue'] = abs(abs(var_diff_rc['RC']) - abs(var_diff_rc['RC_ps']))
    var_summary['DiffRC'] = var_summary['Stem'].map(var_diff_rc.groupby('Stem')['DiffValue'].mean()).fillna(0)

    # ----- Combine -----
    cons_df, var_df = cons_summary.reset_index(drop=True), var_summary.reset_index(drop=True)
    max_len = max(len(cons_df), len(var_df))
    cons_df, var_df = cons_df.reindex(range(max_len)), var_df.reindex(range(max_len))

    summary = pd.DataFrame(index=range(max_len))
    summary['ConstraintStem'] = cons_df['Stem']
    summary['R'] = cons_df['R']
    summary['R_ps'] = cons_df['R_ps']
    summary['#RemovedR'] = cons_df['Removed']
    summary['%RemovedR'] = cons_df['%RemovedR']
    summary['Bound'] = cons_df['Bound']
    summary['Bound_ps'] = cons_df['Bound_ps']
    summary['%BoundReduction'] = cons_df['%BoundReduction']
    summary['#Active'] = cons_df['#Active']
    summary['#Dual'] = cons_df['#Dual']
    summary['#Active_ps'] = cons_df['#Active_ps']
    summary['#Dual_ps'] = cons_df['#Dual_ps']
    summary['#DiffDual'] = cons_df['#DiffDual']
    summary['Dual'] = cons_df['Dual']
    summary['Dual_ps'] = cons_df['Dual_ps']
    summary['DiffDual'] = cons_df['DiffDual']
    # summary['Special'] = cons_df['Special']

    summary[''] = ''
    summary['VariableStem'] = var_df['Stem']
    summary['C'] = var_df['C']
    summary['C_ps'] = var_df['C_ps']
    summary['#RemovedC'] = var_df['Removed']
    summary['%RemovedC'] = var_df['%RemovedC']
    summary['Range'] = var_df['Range']
    summary['Range_ps'] = var_df['Range_ps']
    summary['%RangeReduction'] = var_df['%RangeReduction']
    summary['#RC'] = var_df['#RC']
    summary['#RC_ps'] = var_df['#RC_ps']
    summary['#DiffRC'] = var_df['#DiffRC']
    summary['RC'] = var_df['RC']
    summary['RC_ps'] = var_df['RC_ps']
    summary['DiffRC'] = var_df['DiffRC']
    summary['#LB'] = var_df['#LB']
    summary['#UB'] = var_df['#UB']
    summary['#LB_ps'] = var_df['#LB_ps']
    summary['#UB_ps'] = var_df['#UB_ps']
    summary['%Zero'] = var_df['%Zero']
    summary['%Zero_ps'] = var_df['%Zero_ps']
    # summary['SuperfluousNB'] = var_df['SuperfluousNB']   # <-- commented
    # summary['Superfluous'] = var_df['Superfluous']
    # summary['C_ps/RC=0'] = var_df['C_ps/RC=0']
    # summary['C_ps/RC>0'] = var_df['C_ps/RC>0']

    sense_counts = df_cons['Sense'].value_counts()
    sense_ps_counts = df_cons['Sense_ps'].value_counts()

    def safe_ratio(u, t):
        return 1 - u / t if t else 0

    totals_data = {
        'Type': ['', '', '<', '=', '>', ''],
        'R': [
            None, None,
            sense_counts.get('<', 0),
            sense_counts.get('=', 0),
            sense_counts.get('>', 0),
            None,
        ],
        'R_ps': [
            None, None,
            sense_ps_counts.get('<', 0),
            sense_ps_counts.get('=', 0),
            sense_ps_counts.get('>', 0),
            None,
        ],
    }
    totals_df = pd.DataFrame(totals_data)
    totals_df['#RemovedRow'] = totals_df.apply(lambda r: (r['R'] or 0) - (r['R_ps'] or 0) if pd.notna(r['R']) else None, axis=1)
    totals_df['%RemovedRow'] = totals_df.apply(lambda r: safe_ratio(r['R_ps'], r['R']) if pd.notna(r['R']) else None, axis=1)

    var_block = pd.DataFrame({
        'Type': ['', '', 'C'],
        'C': [None, None, (df_vars['Type'] == 'C').sum()],
        'C_ps': [None, None, (df_vars['Type_ps'] == 'C').sum()],
    })
    var_block['#RemovedCol'] = var_block.apply(lambda r: (r['C'] or 0) - (r['C_ps'] or 0) if pd.notna(r['C']) else None, axis=1)
    var_block['%RemovedCol'] = var_block.apply(lambda r: safe_ratio(r['C_ps'], r['C']) if pd.notna(r['C']) else None, axis=1)

    # ----- Write Excel -----
    print('Writing summary CSV file...')
    output_path = os.path.join(dir_name, f'{case_name}{SUMMARY_SUFFIX}')
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        summary.to_excel(writer, index=False, sheet_name='Summary')
        ws = writer.sheets['Summary']

        # find last used column in the summary sheet
        last_col = summary.shape[1]  # number of summary columns
        start_col = last_col + 2     # +1 for blank, +1 to start writing

        # --- Totals block ---
        for i, col in enumerate(['Type', 'R', 'R_ps', '#RemovedRow', '%RemovedRow']):
            ws.cell(row=1, column=start_col + i, value=col)
        for r_idx, row in totals_df.iterrows():
            for i, col in enumerate(['Type', 'R', 'R_ps', '#RemovedRow', '%RemovedRow']):
                ws.cell(row=r_idx + 2, column=start_col + i, value=row[col])

        # --- Variable type block ---
        start_col2 = start_col + len(['Type', 'R', 'R_ps', '#RemovedRow', '%RemovedRow']) + 1  # leave a blank col
        for i, col in enumerate(['Type', 'C', 'C_ps', '#RemovedCol', '%RemovedCol']):
            ws.cell(row=1, column=start_col2 + i, value=col)
        for r_idx, row in var_block.iterrows():
            for i, col in enumerate(['Type', 'C', 'C_ps', '#RemovedCol', '%RemovedCol']):
                ws.cell(row=r_idx + 2, column=start_col2 + i, value=row[col])

        # Formatting
        apply_percentage_and_colors(ws, ['%RemovedR', '%BoundReduction', '%RemovedC', '%RangeReduction', '%Zero', '%Zero_ps', '%RemovedRow', '%RemovedCol'])
        auto_fit_columns(ws)

    print(f'Summary saved at: {output_path}')


def get_integer_solution_values(m: Model, tol: float = EPSILON) -> dict[str, int]:
    '''
    Extract integer/binary solution values from model `m`.
    Returns {varname: int_value}, rounded with tolerance.
    '''
    vars_ = m.getVars()
    vtypes = m.getAttr('VType', vars_)
    xvals = m.getAttr('X', vars_)

    sol = {}
    for v, t, x in zip(vars_, vtypes, xvals):
        if t in (GRB.BINARY, GRB.INTEGER) and x is not None:
            rounded = int(round(x))
            if abs(x - rounded) <= tol:
                sol[v.VarName] = rounded
    return sol


def fix_model_with_solution_values(src_model: Model, sol: dict[str, int]) -> Model:
    '''
    Return a copy of src_model where integer/binary variables in `sol`
    are fixed to their values and converted to continuous.
    '''
    fixed = src_model.copy()
    vars_ = fixed.getVars()
    vmap = {v.VarName: v for v in vars_}

    for name, value in sol.items():
        if name in vmap:
            v = vmap[name]
            if v.VType in (GRB.BINARY, GRB.INTEGER):
                v.LB = value
                v.UB = value
                v.VType = GRB.CONTINUOUS
    return fixed


def solve_and_fix(m: Model, sol: dict[str, int], label: str):
    '''
    Fix integer vars according to `sol`, re-solve, and return model.
    '''
    fixed = fix_model_with_solution_values(m, sol)
    fixed.optimize()
    fixed.printQuality()
    print('  Kappa ', fixed.KappaExact)
    if fixed.Status == GRB.OPTIMAL:
        print(f'✅ {label} fixed model solved. Obj = {fixed.ObjVal:.6f}')
    else:
        print(f'⚠️ {label} fixed model status {fixed.Status}')
    return fixed


def duals_available(m: Model) -> bool:
    '''
    Check if duals (Pi) are available for constraints in model `m`.
    Requires that the model is optimized and LP-relaxation-like
    (MIP models will not have duals).
    '''
    if m.Status != GRB.OPTIMAL:
        return False
    # At least one constraint must have a non-None Pi
    for c in m.getConstrs():
        if getattr(c, 'Pi', None) is None:
            return False
    return True


# ==============================
# MAIN
# ==============================
def main() -> None:
    file_path = FILE_PATH or os.getcwd()
    output_dir = os.path.join(file_path, f'{CASE_NAME}__Presolve_Info')
    os.makedirs(output_dir, exist_ok=True)

    # ---------- Load original model ----------
    print('🔹 Reading raw model...')
    model_raw = read(f'{CASE_NAME}.lp')

    # ---------- Create presolved model ----------
    print('🔹 Creating presolved model...')
    model_raw.setParam('Presolve', 2)
    model_ps = model_raw.presolve()
    model_ps.write(f'{CASE_NAME}_ps.lp')  # write for inspection

    print('🔹 Solving raw model...')
    model_raw.setParam('MIPGap', 0.01)
    model_raw.setParam('Method', 2)
    model_raw.setParam('Presolve', 2)
    model_raw.setParam('Crossover', -1)
    model_raw.setParam('BarConvTol', 1e-8)
    FileName = os.path.join(output_dir, f'{CASE_NAME}.log')
    if os.path.exists(FileName):
        os.remove(FileName)
    model_raw.setParam('LogFile', FileName)

    # ---------- Solve raw model ----------
    model_raw.optimize()

    sol_vals = {}
    # After solving raw MIP:
    if not duals_available(model_raw):  # no duals because it's a MILP
        sol_vals = get_integer_solution_values(model_raw)
        if sol_vals:
            fixed_lp_raw = fix_model_with_solution_values(model_raw, sol_vals)
            fixed_lp_raw.setParam('Method', 2)
            fixed_lp_raw.setParam('Presolve', 0)
            fixed_lp_raw.setParam('Crossover', -1)
            fixed_lp_raw.setParam('BarConvTol', 1e-8)
            FileName = os.path.join(output_dir, f'{CASE_NAME}_fixed.log')
            if os.path.exists(FileName):
                os.remove(FileName)
            fixed_lp_raw.setParam('LogFile', FileName)
            fixed_lp_raw.optimize()
            fixed_lp_raw.printQuality()
            print('  Kappa ', fixed_lp_raw.KappaExact)
            if duals_available(fixed_lp_raw):
                model_raw_final = fixed_lp_raw
            else:
                # still no duals, pass LP anyway (will be empty)
                model_raw_final = fixed_lp_raw
        else:
            # no integer solution, just solve LP relaxation
            model_raw_lp = model_raw.relax()
            model_raw_lp.setParam('Method', 2)
            model_raw_lp.setParam('Presolve', 0)
            model_raw_lp.setParam('Crossover', -1)
            model_raw_lp.setParam('BarConvTol',1e-8)
            FileName = os.path.join(output_dir, f'{CASE_NAME}_LP.log')
            if os.path.exists(FileName):
                os.remove(FileName)
            model_raw_lp.setParam('LogFile', FileName)
            model_raw_lp.optimize()
            model_raw_lp.printQuality()
            print('  Kappa ', model_raw_lp.KappaExact)
            model_raw_final = model_raw_lp
    else:
        # only for LP problems
        model_raw.printQuality()
        print('  Kappa ', model_raw.KappaExact)
        model_raw_final = model_raw

    # ---------- Solve presolved model ----------
    print('🔹 Solving presolved model...')
    model_ps.setParam('MIPGap', 0.01)
    model_ps.setParam('Method', 2)
    model_ps.setParam('Presolve', 0)
    model_ps.setParam('Crossover', -1)
    model_ps.setParam('BarConvTol', 1e-8)
    FileName = os.path.join(output_dir, f'{CASE_NAME}_ps.log')
    if os.path.exists(FileName):
        os.remove(FileName)
    model_ps.setParam('LogFile', FileName)
    model_ps.optimize()

    sol_vals = {}
    if not duals_available(model_ps):  # no duals because it's a MILP
        sol_vals = get_integer_solution_values(model_ps)
        if sol_vals:
            fixed_lp_ps = fix_model_with_solution_values(model_ps, sol_vals)
            fixed_lp_ps.setParam('Method', 2)
            fixed_lp_ps.setParam('Presolve', 0)
            fixed_lp_ps.setParam('Crossover', -1)
            fixed_lp_ps.setParam('BarConvTol',1e-8)
            FileName = os.path.join(output_dir, f'{CASE_NAME}_ps_fixed.log')
            if os.path.exists(FileName):
                os.remove(FileName)
            fixed_lp_ps.setParam('LogFile', FileName)
            fixed_lp_ps.optimize()
            fixed_lp_ps.printQuality()
            print('  Kappa ', fixed_lp_ps.KappaExact)
            if duals_available(fixed_lp_ps):
                model_ps_final = fixed_lp_ps
            else:
                # still no duals, pass LP anyway (will be empty)
                model_ps_final = fixed_lp_ps
        else:
            # no integer solution, just solve LP relaxation
            model_ps_lp = model_ps.relax()
            model_ps_lp.setParam('Method', 2)
            model_ps_lp.setParam('Presolve', 0)
            model_ps_lp.setParam('Crossover', -1)
            model_ps_lp.setParam('BarConvTol',1e-8)
            FileName = os.path.join(output_dir, f'{CASE_NAME}_ps_LP.log')
            if os.path.exists(FileName):
                os.remove(FileName)
            model_ps_lp.setParam('LogFile', FileName)
            model_ps_lp.optimize()
            model_ps_lp.printQuality()
            print('  Kappa ', model_ps_lp.KappaExact)
            model_ps_final = model_ps_lp
    else:
        # only for LP problems
        model_ps.printQuality()
        print('  Kappa ', model_ps.KappaExact)
        model_ps_final = model_ps

    # ---------- Merge data and write CSVs ----------
    df_vars, df_cons = write_variables_constraints(output_dir, CASE_NAME, model_raw_final, model_ps_final)

    # ---------- Build Excel summary ----------
    build_summary(output_dir, CASE_NAME, df_vars, df_cons)


if __name__ == '__main__':
    main()

